import pdfplumber
import re
from datetime import date, datetime

from openpyxl import load_workbook


class UnsupportedStatementFormatError(Exception):
    """Raised when a PDF does not match any supported statement format."""


PMS_TRANSACTION_PATTERN = re.compile(
    r"^(Buy|Sell)\s+"
    r"(\d{2}/\d{2}/\d{4})\s+"
    r"(\d{2}/\d{2}/\d{4})\s+"
    r"(.+)\s+"
    r"(NSE|BSE)\s+"
    r"([\d,.]+)\s+"
    r"([\d,.]+)\s+"
    r"([\d,.]+)\s+"
    r"([\d,.]+)\s+"
    r"([\d,.]+)$"
)
DATE_PATTERN = re.compile(
    r"(?<!\d)(\d{2}[-/.]\d{2}[-/.]\d{4}|\d{4}-\d{2}-\d{2}|\d{2}-\d{2}-\d{4}|\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}|\d{2}[A-Za-z]{3,9}\d{4})(?!\d)"
)
ICICI_OP_TXN_LINE = re.compile(
    r"^(\d+)\s+"
    r"(\d{2}\.\d{2}\.\d{4})\s+"
    r"(?:(\d+)\s+)?"
    r"([\d,.]+)\s+"
    r"([\d,.]+)$"
)
ICICI_OP_STATEMENT_MARKER = re.compile(
    r"statement of transactions in sav(?:ing|ings)? account",
    re.I,
)
ICICI_OP_TXN_IN_TEXT = re.compile(
    r"\b\d+\s+\d{2}\.\d{2}\.\d{4}\s+(?:\d+\s+)?[\d,.]+\s+[\d,.]+\b"
)
ICICI_OP_SKIP_LINE_MARKERS = (
    "Statement of Transactions",
    "Transaction Withdrawal",
    "S No.",
    "Date Amount",
    "www.icici",
    "Please call",
    "Never share",
    "Legends for",
    "Sincerly",
    "Team ICICI",
    "system generated",
    "RCHG -",
    "PAVC -",
    "SMO -",
    "DTAX -",
    "BPAY -",
    "IDTX -",
    "BBPS -",
    "INFT -",
    "BIL -",
    "ONL -",
    "NEFT -",
    "EBA -",
    "SGB -",
    "PAC -",
    "LNPY -",
    "TOP -",
    "CCWD -",
    "BCTT -",
    "PAYC -",
    "IMPS -",
    "VAT/MAT/NFS -",
    "ATM MMT -",
    "INF -",
    "T Chg -",
    "UCCBRN CMS -",
    "LCCBRN CMS -",
    "N chg -",
    "Fund transfer)",
    "Your Base Branch",
    "Dial your Bank",
)
MONEY_PATTERN = re.compile(r"(?<!\d)\(?-?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d{1,2})?\)?(?!\d)")
TRADE_BOOK_HEADER = re.compile(
    r"Date\s+Stock\s+Action\s+Qty\s+Price\s+Total",
    re.I,
)
TRADE_BOOK_ROW = re.compile(
    r"^(\d{1,2}-[A-Za-z]{3}-\d{2})\s+"
    r"(\S+)\s+"
    r"(Buy|Sell)\s+"
    r"([\d,.]+)\s+"
    r"([\d,.]+)\s+"
    r"([\d,.]+)$",
    re.I,
)
TRADE_BOOK_HEADER_COLUMNS = ("date", "stock", "action", "qty", "price", "total")


def clean_amount(value):
    if value is None:
        return None

    text = str(value).strip()
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"

    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if not cleaned or cleaned == "-":
        return None

    try:
        return float(cleaned)
    except ValueError:
        return None


def is_date_line(line):
    return bool(DATE_PATTERN.search(line.strip()))


MONTHS = {
    "jan": "01",
    "feb": "02",
    "mar": "03",
    "apr": "04",
    "may": "05",
    "jun": "06",
    "jul": "07",
    "aug": "08",
    "sep": "09",
    "sept": "09",
    "oct": "10",
    "nov": "11",
    "dec": "12",
}


def normalize_date(date_string: str) -> str:
    value = date_string.strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        year, month, day = value.split("-")
        return f"{day}-{month}-{year}"

    if re.match(r"^\d{2}-\d{2}-\d{4}$", value):
        day, month, year = value.split("-")
        return f"{day}-{month}-{year}"

    if re.match(r"^\d{2}/\d{2}/\d{4}$", value):
        day, month, year = value.split("/")
        return f"{day}-{month}-{year}"

    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", value):
        day, month, year = value.split(".")
        return f"{day}-{month}-{year}"

    month_match = re.match(r"^(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{4})$", value)
    if month_match:
        day, month_name, year = month_match.groups()
        month = MONTHS.get(month_name[:3].lower())
        if month:
            day = day.zfill(2)
            return f"{day}-{month}-{year}"

    month_match = re.match(r"^(\d{2})([A-Za-z]{3,9})(\d{4})$", value)
    if month_match:
        day, month_name, year = month_match.groups()
        month = MONTHS.get(month_name[:3].lower())
        if month:
            return f"{day}-{month}-{year}"

    month_match = re.match(r"^(\d{1,2})-([A-Za-z]{3})-(\d{2})$", value, re.I)
    if month_match:
        day, month_name, year_short = month_match.groups()
        month = MONTHS.get(month_name[:3].lower())
        if month:
            year = int(year_short)
            if year < 100:
                year += 2000
            return f"{day.zfill(2)}-{month}-{year}"

    return value


def convert_slash_date_to_dmy(date_string: str) -> str:
    day, month, year = date_string.split("/")
    return f"{day}-{month}-{year}"


def infer_transaction_types(transactions, opening_balance=None):
    for index in range(len(transactions)):
        if index == 0:
            if opening_balance is not None:
                if transactions[index]["balance"] > opening_balance:
                    transactions[index]["transaction_type"] = "deposit"
                else:
                    transactions[index]["transaction_type"] = "withdrawal"
            else:
                transactions[index]["transaction_type"] = "unknown"
            continue

        previous_balance = transactions[index - 1]["balance"]
        current_balance = transactions[index]["balance"]

        if current_balance > previous_balance:
            transactions[index]["transaction_type"] = "deposit"
        else:
            transactions[index]["transaction_type"] = "withdrawal"

    return transactions


def extract_money_values(text):
    sanitized = re.sub(DATE_PATTERN, " ", text)
    values = []
    for match in MONEY_PATTERN.finditer(sanitized):
        amount = clean_amount(match.group(0))
        if amount is not None:
            values.append(amount)
    return values


def extract_money_tokens(text):
    sanitized = re.sub(DATE_PATTERN, " ", text)
    return [match.group(0) for match in MONEY_PATTERN.finditer(sanitized)]


def split_line_columns(line):
    columns = re.split(r"\s{2,}|\t", line)
    return [column.strip() for column in columns if column.strip()]


def parse_generic_text_row(line):
    if not line.strip():
        return None

    if re.search(r"\b(total|opening balance|closing balance|b/f|c/f|carry forward|balance brought forward)\b", line, re.I):
        return None

    columns = split_line_columns(line)
    if len(columns) > 1:
        parsed = parse_table_transaction_row(columns)
        if parsed:
            return parsed

    date_match = DATE_PATTERN.search(line)
    if not date_match:
        return None

    amounts = extract_money_values(line)
    if len(amounts) < 2:
        return None

    date_str = date_match.group(0)
    amount = abs(amounts[-2])
    balance = amounts[-1]

    narration = line
    narration = narration.replace(date_str, "", 1)
    for token in extract_money_tokens(line):
        narration = narration.replace(token, " ")
    narration = re.sub(r"\s+", " ", narration).strip(" -|")

    if not narration:
        narration = "Transaction"

    return {
        "date": normalize_date(date_str),
        "narration": narration,
        "amount": amount,
        "balance": balance,
    }


def parse_table_transaction_row(cells):
    row_text = " ".join(cell.strip() for cell in cells if cell)
    if not row_text:
        return None

    if re.search(r"\b(total|opening balance|closing balance|b/f|c/f|carry forward|balance brought forward)\b", row_text, re.I):
        return None

    date_match = None
    for cell in cells:
        if cell and DATE_PATTERN.search(cell):
            date_match = DATE_PATTERN.search(cell)
            break

    if not date_match:
        return None

    money_cells = [clean_amount(cell) for cell in cells if cell and MONEY_PATTERN.search(cell)]
    money_cells = [amount for amount in money_cells if amount is not None]
    if len(money_cells) < 2:
        return None

    balance = money_cells[-1]
    amount = abs(money_cells[-2])
    if len(money_cells) > 2:
        nonzero = [abs(num) for num in money_cells[:-1] if num != 0]
        if nonzero:
            amount = nonzero[-1]

    date_str = date_match.group(0)
    narration_parts = []
    for cell in cells:
        if not cell:
            continue
        if DATE_PATTERN.search(cell) or MONEY_PATTERN.search(cell):
            continue
        narration_parts.append(cell)

    narration = " ".join(narration_parts).strip()
    if not narration:
        narration = row_text
        narration = narration.replace(date_str, "", 1)
        for token in extract_money_tokens(row_text):
            narration = narration.replace(token, " ")
        narration = re.sub(r"\s+", " ", narration).strip(" -|")

    if not narration:
        narration = "Transaction"

    return {
        "date": normalize_date(date_str),
        "narration": narration,
        "amount": amount,
        "balance": balance,
    }


def parse_table_bank_statement(pdf_path):
    transactions = []
    opening_balance = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue

                for row in table:
                    if not any(cell and str(cell).strip() for cell in row):
                        continue

                    cells = [str(cell).strip() if cell is not None else "" for cell in row]
                    parsed = parse_table_transaction_row(cells)
                    if parsed is None:
                        continue

                    if "opening_balance" in parsed:
                        opening_balance = parsed["opening_balance"]
                        continue

                    transactions.append(parsed)

    return infer_transaction_types(transactions, opening_balance)


def parse_generic_transaction_blocks(lines):
    blocks = []
    current_block = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        if is_date_line(line):
            if current_block:
                blocks.append(current_block)
            current_block = [line]
        elif current_block:
            current_block.append(line)

    if current_block:
        blocks.append(current_block)

    return blocks


def build_generic_transaction(block, opening_balance=None):
    if not block:
        return None

    full_text = " ".join(block)
    date_match = DATE_PATTERN.search(full_text)
    if not date_match:
        return None

    full_text_lower = full_text.lower()
    if "opening balance" in full_text_lower or "b/f" in full_text_lower:
        amounts = extract_money_values(full_text)
        if amounts:
            return {"opening_balance": amounts[-1]}
        return None

    amounts = extract_money_values(full_text)
    if len(amounts) < 2:
        return None

    amount = amounts[-2]
    balance = amounts[-1]
    date = normalize_date(date_match.group(0))

    narration = full_text
    narration = narration.replace(date_match.group(0), "", 1)
    for token in extract_money_tokens(full_text):
        narration = narration.replace(token, " ")
    narration = re.sub(r"\s+", " ", narration).strip(" -|")

    if not narration:
        narration = "Transaction"

    return {
        "date": date,
        "narration": narration,
        "amount": amount,
        "balance": balance,
    }


def is_icici_op_transaction_statement(full_text):
    return bool(
        ICICI_OP_STATEMENT_MARKER.search(full_text)
        and ICICI_OP_TXN_IN_TEXT.search(full_text)
    )


def should_skip_icici_op_line(line):
    if not line.strip():
        return True

    if re.match(r"^\d+$", line.strip()):
        return True

    return any(marker in line for marker in ICICI_OP_SKIP_LINE_MARKERS)


def extract_icici_op_transaction_lines(pdf_path):
    lines = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for raw_line in text.split("\n"):
                line = raw_line.strip()
                if should_skip_icici_op_line(line):
                    continue
                lines.append(line)

    return lines


def build_icici_op_transaction(lines, txn_index, txn_indices):
    txn_line_index = txn_indices[txn_index]
    match = ICICI_OP_TXN_LINE.match(lines[txn_line_index])
    if not match:
        return None

    _, date_str, cheque_number, amount_str, balance_str = match.groups()
    next_txn_line_index = (
        txn_indices[txn_index + 1]
        if txn_index + 1 < len(txn_indices)
        else len(lines)
    )

    narration_parts = []
    if txn_index == 0 and txn_line_index > 0:
        narration_parts.append(lines[txn_line_index - 1])

    if txn_index > 0:
        previous_txn_line_index = txn_indices[txn_index - 1]
        previous_between = lines[previous_txn_line_index + 1:txn_line_index]
        if previous_between:
            narration_parts.append(previous_between[-1])

    between_lines = lines[txn_line_index + 1:next_txn_line_index]
    if between_lines:
        if txn_index + 1 < len(txn_indices):
            narration_parts.extend(between_lines[:-1])
        else:
            narration_parts.extend(
                line for line in between_lines if not should_skip_icici_op_line(line)
            )

    narration = re.sub(r"\s+", " ", " ".join(narration_parts)).strip()
    if cheque_number:
        narration = f"Chq {cheque_number} {narration}".strip()
    if not narration:
        narration = "Transaction"

    return {
        "date": normalize_date(date_str),
        "narration": narration,
        "amount": clean_amount(amount_str),
        "balance": clean_amount(balance_str),
    }


def parse_icici_op_transaction_statement(pdf_path):
    lines = extract_icici_op_transaction_lines(pdf_path)
    txn_indices = [
        index
        for index, line in enumerate(lines)
        if ICICI_OP_TXN_LINE.match(line)
    ]

    transactions = []
    for txn_index in range(len(txn_indices)):
        parsed = build_icici_op_transaction(lines, txn_index, txn_indices)
        if parsed and parsed["amount"] is not None and parsed["balance"] is not None:
            transactions.append(parsed)

    return infer_transaction_types(transactions)


def parse_icici_savings_statement(pdf_path):
    transactions = []
    lines = []
    inside_transactions = False

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for line in text.split("\n"):
                line = line.strip()

                if re.search(r"Statement of Transactions in Sav(?:ing|ings) Account", line):
                    inside_transactions = True
                    continue

                if line.startswith("Total:"):
                    inside_transactions = False
                    break

                if inside_transactions:
                    lines.append(line)

    blocks = parse_generic_transaction_blocks(lines)
    opening_balance = None

    for block in blocks:
        parsed = build_generic_transaction(block)
        if parsed is None:
            continue

        if "opening_balance" in parsed:
            opening_balance = parsed["opening_balance"]
            continue

        transactions.append(parsed)

    return infer_transaction_types(transactions, opening_balance)


def parse_pms_transaction_statement(pdf_path):
    transactions = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for line in text.split("\n"):
                line = line.strip()
                match = PMS_TRANSACTION_PATTERN.match(line)
                if not match:
                    continue

                side, tran_date, settlement_date, security, exchange, quantity, _, _, _, amount = match.groups()

                transactions.append({
                    "date": convert_slash_date_to_dmy(tran_date),
                    "narration": (
                        f"{side} {security.strip()} ({exchange}) "
                        f"Qty {quantity} Settle {settlement_date}"
                    ),
                    "amount": clean_amount(amount),
                    "balance": 0,
                    "transaction_type": "deposit" if side == "Sell" else "withdrawal",
                })

    return transactions


def is_trade_book_statement(full_text):
    return bool(
        TRADE_BOOK_HEADER.search(full_text)
        or TRADE_BOOK_ROW.search(full_text)
    )


def build_trade_book_transaction(date_str, stock, action, qty, price, total):
    amount = clean_amount(total)
    if amount is None:
        return None

    side = action.strip().capitalize()
    if side not in {"Buy", "Sell"}:
        return None

    return {
        "date": normalize_date(str(date_str).strip()),
        "narration": f"{side} {stock.strip()} Qty {qty} @ {price}",
        "amount": amount,
        "balance": 0,
        "transaction_type": "deposit" if side == "Sell" else "withdrawal",
    }


def parse_trade_book_line(line):
    match = TRADE_BOOK_ROW.match(line.strip())
    if not match:
        return None
    return build_trade_book_transaction(*match.groups())


def parse_trade_book_pdf(pdf_path):
    transactions = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for line in text.split("\n"):
                parsed = parse_trade_book_line(line)
                if parsed:
                    transactions.append(parsed)

    return transactions


def normalize_trade_book_header(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def find_trade_book_columns(rows):
    for row_index, row in enumerate(rows):
        headers = [normalize_trade_book_header(cell) for cell in row]
        if not headers:
            continue

        column_map = {}
        for column_name in TRADE_BOOK_HEADER_COLUMNS:
            try:
                column_map[column_name] = headers.index(column_name)
            except ValueError:
                column_map[column_name] = None

        if (
            column_map["date"] is not None
            and column_map["stock"] is not None
            and column_map["action"] is not None
            and column_map["total"] is not None
        ):
            return row_index, column_map

    return None, None


def format_trade_book_date(value):
    if isinstance(value, datetime):
        return normalize_date(value.strftime("%d-%b-%y"))
    if isinstance(value, date):
        return normalize_date(value.strftime("%d-%b-%y"))
    return normalize_date(str(value).strip())


def parse_trade_book_excel(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    header_row, columns = find_trade_book_columns(rows)
    if header_row is None or columns is None:
        return []

    transactions = []
    for row in rows[header_row + 1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        try:
            date_value = row[columns["date"]]
            stock_value = row[columns["stock"]]
            action_value = row[columns["action"]]
            qty_value = row[columns["qty"]] if columns["qty"] is not None else ""
            price_value = row[columns["price"]] if columns["price"] is not None else ""
            total_value = row[columns["total"]]
        except IndexError:
            continue

        if date_value is None or action_value is None or total_value is None:
            continue

        parsed = build_trade_book_transaction(
            format_trade_book_date(date_value),
            str(stock_value or "").strip(),
            str(action_value).strip(),
            str(qty_value or "").strip(),
            str(price_value or "").strip(),
            str(total_value).strip(),
        )
        if parsed:
            transactions.append(parsed)

    return transactions


def parse_excel_statement(file_path):
    if file_path.lower().endswith(".xls") and not file_path.lower().endswith(".xlsx"):
        raise UnsupportedStatementFormatError(
            "Legacy .xls files are not supported. Save the trade book as .xlsx and upload again."
        )

    transactions = parse_trade_book_excel(file_path)
    if transactions:
        return transactions

    raise UnsupportedStatementFormatError(
        "Unsupported Excel format. Supported Excel format: trade book with Date, Stock, Action, Qty, Price, and Total columns."
    )


def parse_generic_bank_statement(pdf_path):
    table_transactions = parse_table_bank_statement(pdf_path)
    if table_transactions:
        return table_transactions

    transactions = []
    opening_balance = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            lines = [line.strip() for line in text.split("\n") if line.strip()]

            for line in lines:
                parsed = parse_generic_text_row(line)
                if parsed is None:
                    continue

                if "opening_balance" in parsed:
                    opening_balance = parsed["opening_balance"]
                    continue

                transactions.append(parsed)

            if not transactions:
                blocks = parse_generic_transaction_blocks(lines)
                for block in blocks:
                    parsed = build_generic_transaction(block)
                    if parsed is None:
                        continue

                    if "opening_balance" in parsed:
                        opening_balance = parsed["opening_balance"]
                        continue

                    transactions.append(parsed)

    return infer_transaction_types(transactions, opening_balance)


def extract_pdf_text(pdf_path) -> str:
    chunks = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                chunks.append(text)
    return "\n".join(chunks)


def parse_pdf_statement(pdf_path):
    full_text = extract_pdf_text(pdf_path)

    if not full_text.strip():
        raise UnsupportedStatementFormatError(
            "Could not read any text from this PDF. The file may be scanned or password protected."
        )

    if is_icici_op_transaction_statement(full_text):
        transactions = parse_icici_op_transaction_statement(pdf_path)
        if transactions:
            return transactions
        raise UnsupportedStatementFormatError(
            "ICICI transaction statement detected, but no transactions could be extracted."
        )

    if re.search(r"statement of transactions", full_text, re.I) or re.search(r"account statement", full_text, re.I):
        transactions = parse_icici_savings_statement(pdf_path)
        if transactions:
            return transactions

    is_pms_statement = (
        "TRANSACTION STATEMENT" in full_text
        or "Care Portfolio Managers" in full_text
    )

    if is_pms_statement or PMS_TRANSACTION_PATTERN.search(full_text):
        transactions = parse_pms_transaction_statement(pdf_path)
        if transactions:
            return transactions
        if is_pms_statement:
            raise UnsupportedStatementFormatError(
                "PMS transaction statement detected, but no transactions could be extracted."
            )

    if is_trade_book_statement(full_text):
        transactions = parse_trade_book_pdf(pdf_path)
        if transactions:
            return transactions
        raise UnsupportedStatementFormatError(
            "Trade book statement detected, but no transactions could be extracted."
        )

    transactions = parse_generic_bank_statement(pdf_path)
    if transactions:
        return transactions

    raise UnsupportedStatementFormatError(
        "Unsupported statement format. Supported formats: ICICI-style account statements, "
        "PMS transaction statements, trade book tables, and other date/narration/amount/balance bank statements."
    )
