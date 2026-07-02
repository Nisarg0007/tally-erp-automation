from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import List, Optional

import pdfplumber
from openpyxl import load_workbook


class UnsupportedTradeBookError(Exception):
    """Raised when a file does not match the supported trade-book formats."""


TRADE_BOOK_HEADER_PATTERN = re.compile(r"Date\s+Stock\s+Action\s+Qty\s+Price\s+Total", re.I)
TRADE_BOOK_ROW_PATTERN = re.compile(
    r"^(\d{1,2}-[A-Za-z]{3}-\d{2})\s+(\S+)\s+(Buy|Sell)\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)$",
    re.I,
)
TRADE_BOOK_COLUMNS = ("date", "stock", "action", "qty", "price", "total")
MONTHS = {
    "jan": "01",
    "feb": "02",
    "mar": "03",
    "apr": "04",
    "may": "05",
    "jun": "06",
    "jul": "07",
    "aug": "08",
    "sep": "09",
    "sept": "09",
    "oct": "10",
    "nov": "11",
    "dec": "12",
}


def _clean_decimal(value: object) -> Optional[Decimal]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _normalize_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d-%b-%y")
    if isinstance(value, date):
        return value.strftime("%d-%b-%y")

    text = str(value or "").strip()
    if not text:
        return ""

    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        year, month, day = text.split("-")
        return f"{day}-{month}-{year}"

    if re.match(r"^\d{2}-\d{2}-\d{4}$", text):
        day, month, year = text.split("-")
        return f"{day}-{month}-{year}"

    if re.match(r"^\d{2}/\d{2}/\d{4}$", text):
        day, month, year = text.split("/")
        return f"{day}-{month}-{year}"

    if re.match(r"^\d{1,2}\s+([A-Za-z]{3,9})\s+\d{4}$", text):
        day, month_name, year = re.match(r"^(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{4})$", text).groups()
        month = MONTHS.get(month_name[:3].lower())
        if month:
            return f"{day.zfill(2)}-{month}-{year}"

    if re.match(r"^\d{1,2}-([A-Za-z]{3})-(\d{2})$", text, re.I):
        day, month_name, year_short = re.match(r"^(\d{1,2})-([A-Za-z]{3})-(\d{2})$", text, re.I).groups()
        month = MONTHS.get(month_name[:3].lower())
        if month:
            year = int(year_short)
            if year < 100:
                year += 2000
            return f"{day.zfill(2)}-{month}-{year}"

    return text


def _tally_date(value: str) -> str:
    normalized = _normalize_date(value)
    if not normalized:
        return ""

    match = re.match(r"^(\d{1,2})-(\d{2})-(\d{4})$", normalized)
    if match:
        day, month, year = match.groups()
        return f"{year}{month}{day}"

    match = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})$", normalized)
    if match:
        day, month, year = match.groups()
        return f"{year}{month.zfill(2)}{day.zfill(2)}"

    match = re.match(r"^(\d{2})([A-Za-z]{3})(\d{2,4})$", normalized, re.I)
    if match:
        day, month, year = match.groups()
        return f"{year}{month}{day}"

    digits = re.findall(r"\d+", normalized)
    if len(digits) >= 3:
        day, month, year = digits[:3]
        if len(year) == 4:
            return f"{year}{month.zfill(2)}{day.zfill(2)}"

    return normalized


def _build_trade_row(date_value: object, stock_code: object, action: object, qty_value: object, price_value: object, total_value: object) -> Optional[dict]:
    if date_value is None or stock_code is None or action is None or total_value is None:
        return None

    action_text = str(action).strip().capitalize()
    if action_text not in {"Buy", "Sell"}:
        return None

    quantity = _clean_decimal(qty_value)
    price = _clean_decimal(price_value)
    total = _clean_decimal(total_value)
    if quantity is None or quantity <= 0 or price is None or price <= 0 or total is None or total <= 0:
        return None

    calculated_trade_value = (quantity * price).quantize(Decimal("0.01"))
    charges = (total - calculated_trade_value).quantize(Decimal("0.01"))

    return {
        "date": _normalize_date(date_value),
        "tally_date": _tally_date(_normalize_date(date_value)),
        "stock_code": str(stock_code).strip(),
        "action": action_text,
        "quantity": quantity,
        "price": price,
        "total_amount": total,
        "calculated_trade_value": calculated_trade_value,
        "charges": charges,
        "party_ledger": "ICICI  Bank  Ltd -  Saving A/c",
        "purchase_ledger": "Equity Investment-Purchase",
        "sales_ledger": "Equity Investment-Sales",
        "charges_ledger": "Brokerage & Charges",
        "status": "Pending",
        "source": "trade_manual",
        "narration": f"{action_text} {str(stock_code).strip()} Qty {quantity} @ {price}",
        "include_charges_in_stock_value": False,
        "charge_posting_mode": "separate",
    }


def parse_trade_book_pdf(pdf_path: str) -> List[dict]:
    rows: List[dict] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                cleaned = line.strip()
                if not cleaned:
                    continue
                if TRADE_BOOK_HEADER_PATTERN.search(cleaned):
                    continue
                match = TRADE_BOOK_ROW_PATTERN.match(cleaned)
                if not match:
                    continue
                parsed = _build_trade_row(*match.groups())
                if parsed:
                    rows.append(parsed)
    return rows


def parse_trade_book_excel(file_path: str) -> List[dict]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    header_row_index: Optional[int] = None
    column_map: dict[str, Optional[int]] = {name: None for name in TRADE_BOOK_COLUMNS}

    for row_index, row in enumerate(rows):
        if not row:
            continue
        headers = [str(cell or "").strip().lower() for cell in row]
        if any(header == "date" for header in headers):
            for column_name in TRADE_BOOK_COLUMNS:
                try:
                    column_map[column_name] = headers.index(column_name)
                except ValueError:
                    column_map[column_name] = None
            header_row_index = row_index
            break

    if header_row_index is None:
        return []

    parsed_rows: List[dict] = []
    for row in rows[header_row_index + 1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        try:
            parsed = _build_trade_row(
                row[column_map["date"]] if column_map["date"] is not None and column_map["date"] < len(row) else None,
                row[column_map["stock"]] if column_map["stock"] is not None and column_map["stock"] < len(row) else None,
                row[column_map["action"]] if column_map["action"] is not None and column_map["action"] < len(row) else None,
                row[column_map["qty"]] if column_map["qty"] is not None and column_map["qty"] < len(row) else None,
                row[column_map["price"]] if column_map["price"] is not None and column_map["price"] < len(row) else None,
                row[column_map["total"]] if column_map["total"] is not None and column_map["total"] < len(row) else None,
            )
        except IndexError:
            continue
        if parsed:
            parsed_rows.append(parsed)

    return parsed_rows


def parse_trade_book_file(file_path: str) -> List[dict]:
    lowered = file_path.lower()
    if lowered.endswith(".pdf"):
        rows = parse_trade_book_pdf(file_path)
        if rows:
            return rows
    elif lowered.endswith((".xlsx", ".xls")):
        if lowered.endswith(".xls"):
            raise UnsupportedTradeBookError("Legacy .xls files are not supported. Please upload .xlsx.")
        rows = parse_trade_book_excel(file_path)
        if rows:
            return rows
    else:
        raise UnsupportedTradeBookError("Unsupported trade-book file. Please upload a PDF or Excel trade book.")

    raise UnsupportedTradeBookError("No trade-book rows could be parsed from the uploaded file.")
