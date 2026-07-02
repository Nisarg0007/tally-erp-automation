from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from typing import List
from uuid import uuid4
import xml.etree.ElementTree as ET
import os
import re
import shutil
from sqlalchemy import text

from database import Base, engine, SessionLocal
from models import Transaction, Ledger
from parser import parse_pdf_statement, parse_excel_statement, UnsupportedStatementFormatError
from schemas import TransactionCreate, TransactionUpdate, TransactionUpdateWithId, BulkTransactionUpdate
from tally_import import read_ledgers
from trade_models import TradeRow
from trade_parser import UnsupportedTradeBookError, parse_trade_book_file
from trade_schemas import TradeRowCreate, TradeRowUpdate, TradeRowUpdateWithId, BulkTradeRowUpdate
from trade_xml_generator import build_trade_stock_master_xml, build_trade_voucher_xml

Base.metadata.create_all(bind=engine)


def ensure_trade_table_columns():
    if engine.dialect.name == "sqlite":
        with engine.begin() as conn:
            result = conn.execute(text("PRAGMA table_info(trade_rows);"))
            columns = [row[1] for row in result]
            if "include_charges_in_stock_value" not in columns:
                conn.execute(text("ALTER TABLE trade_rows ADD COLUMN include_charges_in_stock_value VARCHAR DEFAULT 'No'"))
            if "charge_posting_mode" not in columns:
                conn.execute(text("ALTER TABLE trade_rows ADD COLUMN charge_posting_mode VARCHAR DEFAULT 'separate'"))


ensure_trade_table_columns()


def ensure_transaction_source_column():
    if engine.dialect.name == "sqlite":
        with engine.begin() as conn:
            result = conn.execute(text("PRAGMA table_info(transactions);"))
            columns = [row[1] for row in result]
            if "source" not in columns:
                conn.execute(text("ALTER TABLE transactions ADD COLUMN source VARCHAR"))


ensure_transaction_source_column()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
TALLY_COMPANY_NAME = os.getenv("TALLY_COMPANY_NAME", "AMIT NIKUNJ GANDHI")

os.makedirs(UPLOAD_DIR, exist_ok=True)


VOUCHER_TYPES = ["Receipt", "Payment", "Contra", "Journal"]
STATUS_OPTIONS = ["Pending", "Approved", "Exported"]


def parse_exported_xml(file_path: str) -> List[dict]:
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        transactions = []
        
        for message in root.findall(".//TALLYMESSAGE"):
            voucher = message.find("VOUCHER")
            if voucher is None:
                continue
            
            date_elem = voucher.find("DATE")
            narration_elem = voucher.find("NARRATION")
            vouchertypename_elem = voucher.find("VOUCHERTYPENAME")
            
            date_str = date_elem.text if date_elem is not None else ""
            narration = narration_elem.text if narration_elem is not None else ""
            voucher_type = vouchertypename_elem.text if vouchertypename_elem is not None else "Journal"
            
            entries = voucher.findall("ALLLEDGERENTRIES.LIST")
            debit_ledger = None
            credit_ledger = None
            amount = 0.0
            
            for entry in entries:
                ledger_name_elem = entry.find("LEDGERNAME")
                amount_elem = entry.find("AMOUNT")
                is_deemed_elem = entry.find("ISDEEMEDPOSITIVE")
                
                ledger_name = ledger_name_elem.text if ledger_name_elem is not None else ""
                amount_str = amount_elem.text if amount_elem is not None else "0.00"
                is_deemed = is_deemed_elem.text if is_deemed_elem is not None else "Yes"
                
                try:
                    amount_val = float(amount_str)
                except (ValueError, TypeError):
                    amount_val = 0.0
                
                if is_deemed == "No":
                    debit_ledger = ledger_name
                    amount = abs(amount_val)
                elif is_deemed == "Yes":
                    credit_ledger = ledger_name
            
            if date_str:
                transactions.append({
                    "date": convert_tally_date_to_readable(date_str),
                    "narration": narration,
                    "transaction_type": "deposit" if amount > 0 else "withdrawal" if amount < 0 else "unknown",
                    "amount": amount,
                    "balance": 0,
                    "voucher_type": voucher_type,
                    "debit_ledger": debit_ledger,
                    "credit_ledger": credit_ledger,
                    "final_narration": None,
                })
        
        return transactions
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"XML parsing error: {str(exc)}")


def convert_tally_date_to_readable(date_str: str) -> str:
    date_str = date_str.strip()
    if len(date_str) == 8 and date_str.isdigit():
        year = date_str[0:4]
        month = date_str[4:6]
        day = date_str[6:8]
        return f"{day}-{month}-{year}"
    return date_str


def normalize_date_for_tally(date_string: str) -> str:
    date_string = date_string.strip()
    match = re.match(r"^(\d{2})[\-/](\d{2})[\-/](\d{4})$", date_string)
    if match:
        day, month, year = match.groups()
        return f"{year}{month}{day}"

    match = re.match(r"^(\d{4})[\-/](\d{2})[\-/](\d{2})$", date_string)
    if match:
        year, month, day = match.groups()
        return f"{year}{month}{day}"

    match = re.match(r"^(\d{8})$", date_string)
    if match:
        year = date_string[0:4]
        month = date_string[4:6]
        day = date_string[6:8]
        return f"{year}{month}{day}"

    digits = re.findall(r"\d+", date_string)
    if len(digits) >= 3:
        day, month, year = digits[:3]
        if len(year) == 4:
            return f"{year}{month.zfill(2)}{day.zfill(2)}"

    return ""


def validate_transactions_for_export(transactions: List[Transaction]) -> None:
    errors = []

    for index, transaction in enumerate(transactions, start=1):
        voucher_type = (transaction.voucher_type or "").strip()
        debit_ledger = (transaction.debit_ledger or "").strip()
        credit_ledger = (transaction.credit_ledger or "").strip()
        amount = float(transaction.amount or 0.0)

        if not voucher_type:
            errors.append(f"Transaction {index}: voucher type is required")
        if not debit_ledger:
            errors.append(f"Transaction {index}: debit ledger is required")
        if not credit_ledger:
            errors.append(f"Transaction {index}: credit ledger is required")
        if debit_ledger and credit_ledger and debit_ledger == credit_ledger:
            errors.append(f"Transaction {index}: debit and credit ledger cannot be the same ledger")
        if amount <= 0:
            errors.append(f"Transaction {index}: amount must be greater than zero")

    if errors:
        raise ValueError("; ".join(errors))


def _append_ledger_entry(voucher, ledger_name: str, is_deemed_positive: str, amount: float) -> None:
    entry = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
    ET.SubElement(entry, "LEDGERNAME").text = ledger_name
    ET.SubElement(entry, "GSTCLASS")
    ET.SubElement(entry, "ISDEEMEDPOSITIVE").text = is_deemed_positive
    ET.SubElement(entry, "LEDGERFROMITEM").text = "No"
    ET.SubElement(entry, "REMOVEZEROENTRIES").text = "No"
    ET.SubElement(entry, "ISPARTYLEDGER").text = "No"
    ET.SubElement(entry, "PARTYLEDGERNAME")
    ET.SubElement(entry, "AMOUNT").text = f"{amount:.2f}"


def build_tally_xml(transactions: List[Transaction]) -> bytes:
    validate_transactions_for_export(transactions)
    envelope = ET.Element("ENVELOPE")
    header = ET.SubElement(envelope, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"
    request_desc = ET.SubElement(header, "REQUESTDESC")
    ET.SubElement(request_desc, "REPORTNAME").text = "Vouchers"
    static_variables = ET.SubElement(request_desc, "STATICVARIABLES")
    ET.SubElement(static_variables, "SVCURRENTCOMPANY").text = TALLY_COMPANY_NAME

    body = ET.SubElement(envelope, "BODY")
    import_data = ET.SubElement(body, "IMPORTDATA")
    request_data = ET.SubElement(import_data, "REQUESTDATA")

    for transaction in transactions:
        message = ET.SubElement(request_data, "TALLYMESSAGE")
        message.attrib["xmlns:UDF"] = "TallyUDF"

        voucher_type = (transaction.voucher_type or "Journal").strip()
        debit_ledger = (transaction.debit_ledger or "").strip()
        credit_ledger = (transaction.credit_ledger or "").strip()
        amount = abs(float(transaction.amount or 0.0))

        voucher = ET.SubElement(
            message,
            "VOUCHER",
            {
                "VCHTYPE": voucher_type,
                "ACTION": "Create",
            },
        )

        ET.SubElement(voucher, "DATE").text = normalize_date_for_tally(transaction.date)
        ET.SubElement(voucher, "GUID").text = str(uuid4())
        ET.SubElement(voucher, "NARRATION").text = (
            transaction.final_narration or transaction.narration or ""
        )
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = voucher_type
        ET.SubElement(voucher, "EFFECTIVEDATE").text = normalize_date_for_tally(transaction.date)
        ET.SubElement(voucher, "HASCASHFLOW").text = "No"

        _append_ledger_entry(voucher, debit_ledger, "Yes", -amount)
        _append_ledger_entry(voucher, credit_ledger, "No", amount)

    return ET.tostring(envelope, encoding="utf-8", xml_declaration=True)


def build_transaction_workbook(transactions: List[Transaction]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Approved Transactions"

    headers = [
        "Date",
        "Original Narration",
        "Final Narration",
        "Amount",
        "Voucher Type",
        "Debit Ledger",
        "Credit Ledger",
        "Status",
    ]
    worksheet.append(headers)

    for tx in transactions:
        worksheet.append([
            tx.date,
            tx.narration,
            tx.final_narration or tx.narration,
            tx.amount,
            tx.voucher_type or "",
            tx.debit_ledger or "",
            tx.credit_ledger or "",
            tx.status or "",
        ])

    for index in range(1, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 26

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_ledgers_from_excel(file_path: str) -> List[dict]:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    name_index = None
    group_index = None

    for index, value in enumerate(header):
        if value in ["name", "ledger", "ledger name", "ledgername"]:
            name_index = index
        if value in ["group", "group_name", "group name", "parent"]:
            group_index = index

    if name_index is None:
        name_index = 0

    ledgers = []
    for row in rows[1:]:
        if not row or not row[name_index]:
            continue
        name = str(row[name_index]).strip()
        group_name = ""
        if group_index is not None and len(row) > group_index and row[group_index]:
            group_name = str(row[group_index]).strip()
        ledgers.append({"name": name, "group_name": group_name})

    return ledgers


def save_ledger_records(db, ledgers: List[dict]) -> int:
    imported = 0
    seen_names: set[str] = set()
    for ledger_data in ledgers:
        name = (ledger_data.get("name") or "").strip()
        if not name or name in seen_names:
            continue
        seen_names.add(name)
        db.add(Ledger(name=name, group_name=ledger_data.get("group_name") or ""))
        imported += 1
    return imported


def build_transaction_payload(transaction: Transaction) -> dict:
    return {
        "id": transaction.id,
        "date": transaction.date,
        "narration": transaction.narration,
        "transaction_type": transaction.transaction_type,
        "amount": transaction.amount,
        "balance": transaction.balance,
        "voucher_type": transaction.voucher_type,
        "debit_ledger": transaction.debit_ledger,
        "credit_ledger": transaction.credit_ledger,
        "final_narration": transaction.final_narration,
        "status": transaction.status,
        "source": transaction.source,
    }


def build_trade_row_payload(row: TradeRow) -> dict:
    return {
        "id": row.id,
        "date": row.date,
        "tally_date": row.tally_date,
        "stock_code": row.stock_code,
        "action": row.action,
        "quantity": float(row.quantity) if row.quantity is not None else None,
        "price": float(row.price) if row.price is not None else None,
        "total_amount": float(row.total_amount) if row.total_amount is not None else None,
        "calculated_trade_value": float(row.calculated_trade_value) if row.calculated_trade_value is not None else None,
        "charges": float(row.charges) if row.charges is not None else None,
        "party_ledger": row.party_ledger,
        "purchase_ledger": row.purchase_ledger,
        "sales_ledger": row.sales_ledger,
        "charges_ledger": row.charges_ledger,
        "narration": row.narration,
        "status": row.status,
        "source": row.source,
        "include_charges_in_stock_value": row.include_charges_in_stock_value == "Yes",
        "charge_posting_mode": row.charge_posting_mode,
    }


@app.get("/")
def home():
    return {"status": "running"}


@app.post("/upload-pdf")
async def upload_pdf(file: UploadFile = File(...)):
    file_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    if file.filename.lower().endswith(".xml"):
        transactions = parse_exported_xml(file_path)
    elif file.filename.lower().endswith(".pdf"):
        try:
            transactions = parse_pdf_statement(file_path)
        except UnsupportedStatementFormatError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
    elif file.filename.lower().endswith((".xlsx", ".xls")):
        try:
            transactions = parse_excel_statement(file_path)
        except UnsupportedStatementFormatError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
    else:
        raise HTTPException(status_code=400, detail="File must be PDF, Excel, or XML")

    if not transactions:
        raise HTTPException(
            status_code=400,
            detail="No transactions found in the uploaded file.",
        )

    db = SessionLocal()

    try:
        db.query(Transaction).delete()
        for tx in transactions:
            db_transaction = Transaction(
                date=tx["date"],
                narration=tx["narration"],
                transaction_type=tx["transaction_type"],
                amount=tx["amount"],
                balance=tx.get("balance", 0),
                voucher_type=tx.get("voucher_type"),
                debit_ledger=tx.get("debit_ledger"),
                credit_ledger=tx.get("credit_ledger"),
                final_narration=tx.get("final_narration"),
                source=(
                    "xml"
                    if file.filename.lower().endswith(".xml")
                    else "excel"
                    if file.filename.lower().endswith((".xlsx", ".xls"))
                    else "pdf"
                ),
            )
            db.add(db_transaction)
        db.commit()
        return {"success": True, "count": len(transactions)}
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        db.close()


@app.get("/transactions")
def get_transactions():
    db = SessionLocal()
    try:
        transactions = db.query(Transaction).order_by(Transaction.id).all()
        return [build_transaction_payload(tx) for tx in transactions]
    finally:
        db.close()


@app.put("/transactions/bulk")
def bulk_update_transactions(data: BulkTransactionUpdate):
    db = SessionLocal()
    try:
        updated = 0
        for item in data.transactions:
            transaction = db.query(Transaction).filter(Transaction.id == item.id).first()
            if not transaction:
                continue
            if item.date is not None:
                transaction.date = item.date
            if item.amount is not None:
                transaction.amount = item.amount
            if item.voucher_type is not None:
                transaction.voucher_type = item.voucher_type
            if item.debit_ledger is not None:
                transaction.debit_ledger = item.debit_ledger
            if item.credit_ledger is not None:
                transaction.credit_ledger = item.credit_ledger
            if item.final_narration is not None:
                transaction.final_narration = item.final_narration
            if item.status is not None:
                transaction.status = item.status
            updated += 1
        db.commit()
        return {"success": True, "updated": updated}
    finally:
        db.close()


@app.get("/transactions/{transaction_id}")
def get_transaction(transaction_id: int):
    db = SessionLocal()
    try:
        transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        return build_transaction_payload(transaction)
    finally:
        db.close()


@app.put("/transactions/{transaction_id}")
def update_transaction(transaction_id: int, data: TransactionUpdate):
    db = SessionLocal()
    try:
        transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")

        if data.voucher_type is not None:
            transaction.voucher_type = data.voucher_type
        if data.debit_ledger is not None:
            transaction.debit_ledger = data.debit_ledger
        if data.credit_ledger is not None:
            transaction.credit_ledger = data.credit_ledger
        if data.final_narration is not None:
            transaction.final_narration = data.final_narration
        if data.status is not None:
            transaction.status = data.status

        db.commit()
        return {"success": True}
    finally:
        db.close()


@app.post("/transactions")
def create_transaction(data: TransactionCreate):
    db = SessionLocal()
    try:
        transaction = Transaction(
            date=data.date,
            narration=data.narration,
            transaction_type=data.transaction_type or "Manual",
            amount=data.amount,
            balance=data.balance,
            voucher_type=data.voucher_type,
            debit_ledger=data.debit_ledger,
            credit_ledger=data.credit_ledger,
            final_narration=data.final_narration,
            status=data.status or "Pending",
            source="manual",
        )
        db.add(transaction)
        db.commit()
        db.refresh(transaction)
        return {"success": True, "transaction": build_transaction_payload(transaction)}
    finally:
        db.close()


@app.delete("/transactions/{transaction_id}")
def delete_transaction(transaction_id: int):
    db = SessionLocal()
    try:
        transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        db.delete(transaction)
        db.commit()
        return {"success": True}
    finally:
        db.close()


@app.delete("/transactions")
def clear_transactions():
    db = SessionLocal()
    try:
        db.query(Transaction).delete()
        db.commit()
        return {"success": True}
    finally:
        db.close()


@app.get("/ledgers")
def get_ledgers():
    db = SessionLocal()
    try:
        ledgers = db.query(Ledger).order_by(Ledger.name).all()
        return [{"id": ledger.id, "name": ledger.name, "group_name": ledger.group_name} for ledger in ledgers]
    finally:
        db.close()


@app.post("/import-ledgers")
async def import_ledgers(file: UploadFile = File(...)):
    filename = file.filename.lower()
    if not filename.endswith(".xml") and not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File type must be .xml or .xlsx")

    file_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        if filename.endswith(".xml"):
            ledger_rows = read_ledgers(file_path)
        else:
            ledger_rows = parse_ledgers_from_excel(file_path)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    if not ledger_rows:
        raise HTTPException(status_code=400, detail="No ledgers found in file")

    db = SessionLocal()
    try:
        db.query(Ledger).delete()
        imported = save_ledger_records(db, ledger_rows)
        db.commit()
        return {"success": True, "imported": imported, "total": len(ledger_rows)}
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        db.close()


@app.get("/stats")
def get_stats():
    db = SessionLocal()
    try:
        total = db.query(Transaction).count()
        pending = db.query(Transaction).filter(Transaction.status == "Pending").count()
        approved = db.query(Transaction).filter(Transaction.status == "Approved").count()
        exported = db.query(Transaction).filter(Transaction.status == "Exported").count()
        return {"total": total, "pending": pending, "approved": approved, "exported": exported}
    finally:
        db.close()


@app.post("/trade/import")
async def import_trade_rows(file: UploadFile = File(...)):
    filename = (file.filename or "").lower()
    if not filename.endswith((".pdf", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Trade-book import supports PDF and Excel files only")

    file_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        rows = parse_trade_book_file(file_path)
    except UnsupportedTradeBookError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if not rows:
        raise HTTPException(status_code=400, detail="No trade-book rows were parsed")

    db = SessionLocal()
    try:
        db.query(TradeRow).delete()
        for row in rows:
            db.add(TradeRow(
                date=row["date"],
                tally_date=row["tally_date"],
                stock_code=row["stock_code"],
                action=row["action"],
                quantity=row["quantity"],
                price=row["price"],
                total_amount=row["total_amount"],
                calculated_trade_value=row["calculated_trade_value"],
                charges=row["charges"],
                party_ledger=row["party_ledger"],
                purchase_ledger=row["purchase_ledger"],
                sales_ledger=row["sales_ledger"],
                charges_ledger=row["charges_ledger"],
                narration=row["narration"],
                status=row["status"],
                source=row["source"],
                include_charges_in_stock_value="Yes" if row.get("include_charges_in_stock_value") else "No",
                charge_posting_mode=row.get("charge_posting_mode", "separate"),
            ))
        db.commit()
        return {"success": True, "count": len(rows)}
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    finally:
        db.close()


@app.get("/trade/rows")
def get_trade_rows():
    db = SessionLocal()
    try:
        rows = db.query(TradeRow).order_by(TradeRow.id).all()
        return [build_trade_row_payload(row) for row in rows]
    finally:
        db.close()


@app.put("/trade/rows/bulk")
def bulk_update_trade_rows(data: BulkTradeRowUpdate):
    db = SessionLocal()
    try:
        updated = 0
        for item in data.rows:
            row = db.query(TradeRow).filter(TradeRow.id == item.id).first()
            if not row:
                continue
            for field in [
                "date", "stock_code", "action", "quantity", "price", "total_amount",
                "calculated_trade_value", "charges", "party_ledger", "purchase_ledger",
                "sales_ledger", "charges_ledger", "narration", "status"
            ]:
                value = getattr(item, field, None)
                if value is not None:
                    setattr(row, field, value)
            if item.include_charges_in_stock_value is not None:
                row.include_charges_in_stock_value = "Yes" if item.include_charges_in_stock_value else "No"
            if item.charge_posting_mode is not None:
                row.charge_posting_mode = item.charge_posting_mode
            updated += 1
        db.commit()
        return {"success": True, "updated": updated}
    finally:
        db.close()


@app.post("/trade/rows")
def create_trade_row(data: TradeRowCreate):
    db = SessionLocal()
    try:
        row = TradeRow(
            date=data.date,
            tally_date="",
            stock_code=data.stock_code,
            action=data.action,
            quantity=data.quantity,
            price=data.price,
            total_amount=data.total_amount,
            calculated_trade_value=data.calculated_trade_value,
            charges=data.charges,
            party_ledger=data.party_ledger,
            purchase_ledger=data.purchase_ledger,
            sales_ledger=data.sales_ledger,
            charges_ledger=data.charges_ledger,
            narration=data.narration,
            status=data.status or "Pending",
            source=data.source or "manual",
            include_charges_in_stock_value="Yes" if data.include_charges_in_stock_value else "No",
            charge_posting_mode=data.charge_posting_mode or "separate",
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        return {"success": True, "row": build_trade_row_payload(row)}
    finally:
        db.close()


@app.put("/trade/rows/{row_id}")
def update_trade_row(row_id: int, data: TradeRowUpdate):
    db = SessionLocal()
    try:
        row = db.query(TradeRow).filter(TradeRow.id == row_id).first()
        if not row:
            raise HTTPException(status_code=404, detail="Trade row not found")
        for field in [
            "date", "stock_code", "action", "quantity", "price", "total_amount",
            "calculated_trade_value", "charges", "party_ledger", "purchase_ledger",
            "sales_ledger", "charges_ledger", "narration", "status"
        ]:
            value = getattr(data, field, None)
            if value is not None:
                setattr(row, field, value)
        if data.include_charges_in_stock_value is not None:
            row.include_charges_in_stock_value = "Yes" if data.include_charges_in_stock_value else "No"
        if data.charge_posting_mode is not None:
            row.charge_posting_mode = data.charge_posting_mode
        db.commit()
        return {"success": True}
    finally:
        db.close()


@app.delete("/trade/rows/{row_id}")
def delete_trade_row(row_id: int):
    db = SessionLocal()
    try:
        row = db.query(TradeRow).filter(TradeRow.id == row_id).first()
        if not row:
            raise HTTPException(status_code=404, detail="Trade row not found")
        db.delete(row)
        db.commit()
        return {"success": True}
    finally:
        db.close()


@app.delete("/trade/rows")
def clear_trade_rows():
    db = SessionLocal()
    try:
        db.query(TradeRow).delete()
        db.commit()
        return {"success": True}
    finally:
        db.close()


@app.post("/trade/validate-stock-masters")
def validate_trade_stock_masters():
    db = SessionLocal()
    try:
        rows = db.query(TradeRow).filter(TradeRow.status == "Approved").all()
        stock_codes = sorted({row.stock_code for row in rows if row.stock_code})
        missing = []
        for stock_code in stock_codes:
            if not stock_code:
                continue
            existing = db.query(Ledger).filter(Ledger.name == stock_code).first()
            if existing is None:
                missing.append(stock_code)
        return {"success": True, "stock_items": stock_codes, "missing": missing}
    finally:
        db.close()


@app.post("/trade/export")
def export_trade_vouchers():
    db = SessionLocal()
    try:
        rows = db.query(TradeRow).filter(TradeRow.status == "Approved").order_by(TradeRow.id).all()
        if not rows:
            raise HTTPException(status_code=404, detail="No approved trade rows available for export")

        trade_rows = []
        for row in rows:
            if not row.stock_code:
                continue
            quantity = row.quantity or Decimal("0")
            price = row.price or Decimal("0")
            total_amount = row.total_amount or Decimal("0")
            if quantity <= 0 or price <= 0 or total_amount <= 0:
                continue
            trade_rows.append({
                "action": row.action or "Buy",
                "stock_code": row.stock_code,
                "quantity": quantity,
                "price": price,
                "total_amount": total_amount,
                "tally_date": row.tally_date or row.date,
                "narration": row.narration or f"{row.action} {row.stock_code}",
                "party_ledger": row.party_ledger or "ICICI  Bank  Ltd -  Saving A/c",
                "charges_ledger": row.charges_ledger or "Brokerage & Charges",
                "status": row.status,
                "charge_posting_mode": row.charge_posting_mode or "separate",
            })
        payload = build_trade_voucher_xml(trade_rows)
        return StreamingResponse(
            BytesIO(payload),
            media_type="application/xml",
            headers={"Content-Disposition": "attachment; filename=trade_vouchers.xml"},
        )
    finally:
        db.close()


@app.post("/trade/export-stock-masters")
def export_trade_stock_masters():
    db = SessionLocal()
    try:
        rows = db.query(TradeRow).filter(TradeRow.status == "Approved").all()
        stock_codes = sorted({row.stock_code for row in rows if row.stock_code})
        payload = build_trade_stock_master_xml([{"name": name, "group": "Equity Shares"} for name in stock_codes])
        return StreamingResponse(
            BytesIO(payload),
            media_type="application/xml",
            headers={"Content-Disposition": "attachment; filename=trade_stock_masters.xml"},
        )
    finally:
        db.close()


@app.get("/export/tally-xml")
def export_tally_xml():
    db = SessionLocal()
    try:
        transactions = db.query(Transaction).filter(Transaction.status == "Approved").order_by(Transaction.id).all()
        if not transactions:
            raise HTTPException(status_code=404, detail="No approved transactions available for export")

        try:
            payload = build_tally_xml(transactions)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

        return StreamingResponse(
            BytesIO(payload),
            media_type="application/xml",
            headers={"Content-Disposition": "attachment; filename=tally_vouchers.xml"},
        )
    finally:
        db.close()


@app.get("/export/transactions.xlsx")
def export_transactions_xlsx():
    db = SessionLocal()
    try:
        transactions = db.query(Transaction).filter(Transaction.status == "Approved").order_by(Transaction.id).all()
        if not transactions:
            raise HTTPException(status_code=404, detail="No approved transactions available for export")

        workbook = build_transaction_workbook(transactions)
        return StreamingResponse(
            workbook,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=approved_transactions.xlsx"},
        )
    finally:
        db.close()

        db.close()